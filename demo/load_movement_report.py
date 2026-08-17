#!/usr/bin/env python3
"""
Ingest RORC Item Movement Reports and turn them into observed demand.

The price book tells us what the store carries and what it costs. It says
nothing about what actually sells. Item Movement Reports do - they carry real
units, real dollars, and real cost per UPC for a date range.

Anything covered by a movement report stops being modelled and becomes
measured. Feed in more reports and more of the demo becomes real data.

Usage:
    python3 load_movement_report.py report.xlsx [more.xlsx ...]
    python3 load_movement_report.py reports/*.xlsx

Output:
    movement_actuals.csv  - one row per UPC per report, with a daily rate

Then rebuild:
    python3 make_sample_catalogue.py
    python3 generate_demo_data.py
"""

import csv
import os
import re
import sys
import glob
import collections
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("This needs openpyxl:  pip3 install openpyxl")
    sys.exit(1)

OUTPUT = "movement_actuals.csv"

# Header labels vary slightly between report templates, so each field is
# matched against a list of aliases rather than a fixed column position.
FIELD_ALIASES = {
    "upc":         ["upc"],
    "description": ["description"],
    "brand":       ["brand"],
    "department":  ["major dept.", "major dept", "department"],
    "sub_dept":    ["e. dept.", "e. dept", "sub dept"],
    "category":    ["category"],
    "vendor_name": ["vendor name"],
    "qty_sold":    ["qty sold", "qty  sold", "quantity sold"],
    "sales":       ["sales"],
    "ext_cost":    ["extended cost", "ext cost"],
    "cost_unit":   ["cost/unit", "cost per unit", "unit cost"],
    "gm_pct":      ["gm %", "gm%"],
    "gm_dollars":  ["gm $", "gm$"],
    "price":       ["price"],
    "store":       ["store"],
    "last_sold":   ["last sold"],
    "size":        ["size"],
    "pack":        ["pack"],
}

SKIP_ROW_MARKERS = ("subtotal", "grand total", "total :", "store  name",
                    "store name", "selected  driver", "selected driver",
                    "date range")


def norm(s):
    return re.sub(r"\s+", " ", str(s or "").strip().lower())


def parse_date_range(cells):
    """
    Pull the reporting window out of a header line like
    "Date Range: 8/7/2026-8/14/2026".
    """
    blob = " ".join(str(c) for c in cells if c)
    m = re.search(r"(\d{1,2}/\d{1,2}/\d{2,4})\s*-\s*(\d{1,2}/\d{1,2}/\d{2,4})", blob)
    if not m:
        return None, None
    out = []
    for raw in m.groups():
        for fmt in ("%m/%d/%Y", "%m/%d/%y"):
            try:
                out.append(datetime.strptime(raw, fmt).date())
                break
            except ValueError:
                continue
    return (out[0], out[1]) if len(out) == 2 else (None, None)


def parse_store(cells):
    """Read "Store Name: 000001 (Riverside)" into a code and a name."""
    blob = " ".join(str(c) for c in cells if c)
    m = re.search(r"store\s*name\s*:\s*(\S+)\s*(?:\(([^)]*)\))?", blob, re.I)
    if not m:
        return None, None
    code = m.group(1).strip()
    name = (m.group(2) or "").strip()
    if code.isdigit():
        code = f"STORE{int(code):02d}"
    return code, name


def num(v):
    if v is None:
        return None
    s = re.sub(r"[^0-9.\-]", "", str(v))
    if s in ("", ".", "-"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def find_header(ws):
    """Locate the header row and map our field names onto column indexes."""
    for r_i, row in enumerate(ws.iter_rows(min_row=1, max_row=12, values_only=True), 1):
        labels = {norm(c): i for i, c in enumerate(row) if c}
        if "description" in labels and any(
                a in labels for a in FIELD_ALIASES["qty_sold"]):
            mapping = {}
            for field, aliases in FIELD_ALIASES.items():
                for a in aliases:
                    if a in labels:
                        mapping[field] = labels[a]
                        break
            return r_i, mapping
    return None, None


def parse_workbook(path):
    # Not read_only: these exports carry cached formula values that the
    # read-only reader reports as an empty sheet.
    wb = openpyxl.load_workbook(path, data_only=True)
    out = []

    for ws in wb.worksheets:
        header_row, cols = find_header(ws)
        if not cols:
            continue

        rows = list(ws.iter_rows(values_only=True))
        start, end = None, None
        store_code, store_name = None, None

        # Header block sits above the data and carries the window and store
        for row in rows[:header_row + 6]:
            joined = norm(" ".join(str(c) for c in row if c))
            if "date range" in joined and not start:
                start, end = parse_date_range(row)
            if "store" in joined and "name" in joined and not store_code:
                store_code, store_name = parse_store(row)

        days = ((end - start).days + 1) if (start and end) else None

        def cell(row, field):
            i = cols.get(field)
            return row[i] if (i is not None and i < len(row)) else None

        for row in rows[header_row:]:
            if not row or all(c is None for c in row):
                continue
            joined = norm(" ".join(str(c) for c in row if c))
            if any(joined.startswith(m) or m in joined[:30]
                   for m in SKIP_ROW_MARKERS):
                continue

            desc = str(cell(row, "description") or "").strip()
            upc_raw = cell(row, "upc")
            qty = num(cell(row, "qty_sold"))

            if not desc or qty is None:
                continue

            upc = re.sub(r"[^0-9]", "", str(upc_raw or ""))
            if not upc:
                continue

            sales = num(cell(row, "sales")) or 0.0
            cost_u = num(cell(row, "cost_unit"))
            price = num(cell(row, "price"))
            gm_d = num(cell(row, "gm_dollars"))

            # Category arrives as "00110111 - BERRIES"
            cat_raw = str(cell(row, "category") or "").strip()
            cm = re.match(r"(\d+)\s*-\s*(.+)", cat_raw)
            cat_code, cat_name = (cm.group(1), cm.group(2).strip()) if cm \
                else ("", cat_raw)

            out.append({
                "upc":             upc,
                "item_id":         upc.lstrip("0").rjust(14, "0"),
                "item_desc":       desc,
                "brand":           str(cell(row, "brand") or "").strip(),
                "department_code": str(cell(row, "department") or "").strip(),
                "category_code":   cat_code,
                "category_name":   cat_name,
                "vendor":          str(cell(row, "vendor_name") or "").strip(),
                "store_code":      store_code or "",
                "store_name":      store_name or "",
                "period_start":    start.isoformat() if start else "",
                "period_end":      end.isoformat() if end else "",
                "days":            days or "",
                "units":           qty,
                "sales":           round(sales, 2),
                "unit_price":      round(price, 2) if price else
                                   (round(sales / qty, 2) if qty else 0),
                "unit_cost":       round(cost_u, 4) if cost_u else "",
                "gross_margin":    round(gm_d, 2) if gm_d is not None else "",
                "units_per_day":   round(qty / days, 4) if days else "",
                "sales_per_day":   round(sales / days, 4) if days else "",
                "source_file":     os.path.basename(path),
            })

    return out


def main():
    args = sys.argv[1:]
    if not args:
        found = sorted(glob.glob("*.xlsx")) + sorted(glob.glob("reports/*.xlsx"))
        if not found:
            print(__doc__)
            sys.exit(1)
        args = found

    paths = []
    for a in args:
        paths.extend(sorted(glob.glob(a)) if any(c in a for c in "*?[") else [a])

    all_rows = []
    for p in paths:
        if not os.path.exists(p):
            print(f"  skipping {p} - not found")
            continue
        rows = parse_workbook(p)
        print(f"  {os.path.basename(p)}: {len(rows)} item rows")
        all_rows.extend(rows)

    if not all_rows:
        print("Nothing parsed. Is this an Item Movement Report export?")
        sys.exit(1)

    # If the same UPC shows up in several reports, keep the most recent window
    best = {}
    for r in all_rows:
        key = (r["item_id"], r["store_code"])
        prev = best.get(key)
        if not prev or (r["period_end"] or "") > (prev["period_end"] or ""):
            best[key] = r
    rows = list(best.values())

    fields = ["upc", "item_id", "item_desc", "brand", "department_code",
              "category_code", "category_name", "vendor", "store_code",
              "store_name", "period_start", "period_end", "days", "units",
              "sales", "unit_price", "unit_cost", "gross_margin",
              "units_per_day", "sales_per_day", "source_file"]

    with open(OUTPUT, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(rows)

    total_u = sum(r["units"] for r in rows)
    total_s = sum(r["sales"] for r in rows)
    cats = collections.Counter(r["category_name"] for r in rows)
    stores = collections.Counter(r["store_code"] for r in rows if r["store_code"])
    windows = sorted({(r["period_start"], r["period_end"]) for r in rows})

    print()
    print(f"Wrote {OUTPUT}")
    print(f"  items with observed movement: {len(rows):,}")
    print(f"  total units:                  {total_u:,.0f}")
    print(f"  total sales:                  ${total_s:,.2f}")
    print(f"  stores:                       {', '.join(stores) or 'unknown'}")
    for s, e in windows:
        print(f"  window:                       {s} to {e}")
    print()
    print("  categories covered:")
    for c, n in cats.most_common(10):
        print(f"    {c or '(none)'}: {n} items")
    print()
    print("Next:")
    print("  python3 make_sample_catalogue.py")
    print("  python3 generate_demo_data.py")


if __name__ == "__main__":
    main()
